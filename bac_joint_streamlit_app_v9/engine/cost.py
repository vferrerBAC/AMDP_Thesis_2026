"""
engine/cost.py  —  IJET Milestone 2 cost analysis (Option 1: Excel-as-engine)

This module does NOT reimplement any cost formula. It keeps Steve Rosario's
BAC cost calculator workbook as the calculation engine, exactly as authored,
and only (a) writes IJET's inputs into the template's input columns, (b) lets
Excel recalculate Steve's formulas natively via COM, and (c) reads the results
back. Not one formula is retyped, so the logic is preserved byte-for-byte.

ENVIRONMENT REQUIREMENT
    Windows with desktop Microsoft Excel installed + pywin32.
    (Same COM approach IJET already uses for Autodesk Inventor. Cannot run in a
    cloud container — this belongs on the local-desktop side of the pipeline.)

FAITHFULNESS MECHANISM
    Inputs are appended into the workbook's real Excel Tables (PartsListTable,
    JointsListTable, SummaryTable) via ListRows.Add. Excel auto-fills every
    calculated column of Steve's formulas into each new row. We never write a
    formula; we only set the input cells.

INPUT CONTRACT  (produced upstream by Block 1 extraction / an adapter)
    parts: list[dict], each with keys ->
        part_set            str    assembly / set name  (BAC Part List col A, and Summary Set ID)
        part_identifier     str    (col B)
        part_quantity       int    (col C)
        part_class          str    "sheet_metal" | "tube"  (only used to pick a default cutting method)
        cutting_method      str|None  (col D; default applied if None)
        forming_method      str|None  (col E; default applied if None)
        ncx_material        str    (col F)
        gauge               float  (col G)
        pierce_count        float  (col H)
        cut_distance_inches float  (col I)
        unique_bends        float  (col J)
        corner_weld         float  (col K)
        flat_length_inches  float  (col L)
        flat_width_inches   float  (col M)
        assembly_category   str    (col N)
    joints: list[dict], each with keys ->
        part_a              str    (col A)
        part_b              str    (col B)
        joint_length_inches float  (col C)
        (Weld Check is intentionally NOT supplied — the template computes col O
         itself. This resolves the V4-macro vs. template drift in favor of the
         template's own formula.)

OUT OF SCOPE (mirrors what Steve's macro touched)
    Purchase Part List and standalone Hardware sheets are not populated here, so
    the corresponding Summary columns read 0. Structural (BAC) parts + joints
    only, consistent with the V4 macro.

FIRST-RUN VALIDATION (the three spots most likely to need a tweak on real data)
    1. Table append behavior — clearing/adding ListRows on PartsListTable /
       JointsListTable; confirm calculated columns auto-fill as expected.
    2. Summary Set-ID rows — one SummaryTable row per distinct part_set.
    3. Output cell mapping in _read_outputs — confirm the header names match.
    Validate by running one assembly through this and diffing the Summary totals
    against Steve's macro output for the same assembly.
"""

from __future__ import annotations

import os
import shutil
from datetime import datetime

# --- Excel enum literals (avoids gencache / EnsureDispatch dependency) ---
XL_CALC_MANUAL = -4135
XL_CALC_AUTOMATIC = -4105

# --- Workbook input geometry (1-based column indices within each table) ---
# BAC Part List: header row 3, first data row 4, table starts at column A.
PART_COL = {
    "part_set": 1, "part_identifier": 2, "part_quantity": 3,
    "cutting_method": 4, "forming_method": 5, "ncx_material": 6, "gauge": 7,
    "pierce_count": 8, "cut_distance_inches": 9, "unique_bends": 10,
    "corner_weld": 11, "flat_length_inches": 12, "flat_width_inches": 13,
    "assembly_category": 14,
}
JOINT_COL = {"part_a": 1, "part_b": 2, "joint_length_inches": 3}

# Fields that MUST be present. Missing/unparseable -> needs_review (never a silent 1).
REQUIRED_PART_FIELDS = [
    "part_set", "part_identifier", "part_quantity", "ncx_material", "gauge",
    "pierce_count", "cut_distance_inches", "unique_bends", "corner_weld",
    "flat_length_inches", "flat_width_inches", "assembly_category",
]
REQUIRED_JOINT_FIELDS = ["part_a", "part_b", "joint_length_inches"]

# --- Method presets ---------------------------------------------------------
# Forming: Manual Press Brake is the legacy process (Valeria's call).
FORMING_METHOD_DEFAULT = "Manual Press Brake"
# Cutting: tube parts go on the tube laser (physically correct + ties to M3).
CUTTING_METHOD_DEFAULT_TUBE = "Auto Tube Laser"
# Cutting for sheet-metal parts: CONFIRM WITH STEVE which manual cut is legacy.
# Change this single line if the legacy sheet-metal cutting process differs.
CUTTING_METHOD_DEFAULT_SHEET_METAL = "Manual Shear Punch"

# --- Workbook-level inputs ---
DEFAULT_REGION = "NA"          # BAC Part List!D1  (options: CN | EU | NA)
DEFAULT_LENGTH_UOM = "in"      # BAC Part List!F1  (options: in | mm)

# Sheet / table names in the template
SHEET_PARTS, TABLE_PARTS = "BAC Part List", "PartsListTable"
SHEET_JOINTS, TABLE_JOINTS = "Joints List", "JointsListTable"
SHEET_SUMMARY, TABLE_SUMMARY = "Summary", "SummaryTable"

# Per-part cost columns surfaced back to the app after a run. Selected by header
# name (not position) so column reordering in the template can't break the read.
PART_COST_COLUMNS = [
    "Part Identifier", "Part Quantity",
    "Unit Material Cost", "Unit Fully Burdened Labor Cost", "Unit Fully Burdened Total Cost",
    "Extended Material Cost", "Extended Fully Burdened Labor Cost", "Extended Fully Burdened Total Cost",
]


# ---------------------------------------------------------------------------
# Public: dropdown options (lightweight, openpyxl only — no Excel/COM needed)
# ---------------------------------------------------------------------------
def read_method_options(template_path: str) -> dict:
    """Return the option lists for the Streamlit dropdowns, read straight from
    the template's named ranges so the UI never drifts from Steve's lists.
    Safe to call without Excel installed (used by the app layer)."""
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries

    wb = load_workbook(template_path, data_only=False, read_only=True)

    def values(defined_name: str):
        dn = wb.defined_names[defined_name]
        out = []
        for sheet_name, coord in dn.destinations:
            ws = wb[sheet_name]
            min_c, min_r, max_c, max_r = range_boundaries(coord.replace("$", ""))
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    v = ws.cell(r, c).value
                    if v not in (None, ""):
                        out.append(v)
        return out

    opts = {
        "cutting_method": values("CuttingMethod"),
        "forming_method": values("FormingMethod"),
        "fabrication_region": values("FabricationRegion"),
        "length_uom": values("LengthUnitofMeasure"),
    }
    wb.close()
    return opts


# ---------------------------------------------------------------------------
# Public: run the cost analysis
# ---------------------------------------------------------------------------
def run_cost_analysis(
    parts: list[dict],
    joints: list[dict],
    template_path: str,
    output_dir: str,
    region: str = DEFAULT_REGION,
    length_uom: str = DEFAULT_LENGTH_UOM,
) -> dict:
    """
    Populate a fresh copy of Steve's template with IJET inputs, recalc via Excel
    COM, and read the Summary rollups back.

    Returns:
        {
          "output_workbook": <path to the saved per-run workbook>,
          "summary": [ {header: value, ...}, ... ],   # SummaryTable rows
          "parts": [ {Part Identifier, ..., cost columns}, ... ],  # per-part costs
          "needs_review": [ {scope, identifier, field, issue}, ... ],
        }
    """
    import win32com.client as win32  # local import: only needed on Windows

    needs_review: list[dict] = []
    clean_parts = [_prepare_part(p, needs_review) for p in parts]
    clean_joints = [_prepare_joint(j, needs_review) for j in joints]

    working_path = _stage_workbook(template_path, output_dir)

    excel = None
    wb = None
    try:
        excel = win32.DispatchEx("Excel.Application")  # fresh, isolated instance
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        wb = excel.Workbooks.Open(os.path.abspath(working_path))
        excel.Calculation = XL_CALC_MANUAL

        ws_parts = wb.Sheets(SHEET_PARTS)
        ws_joints = wb.Sheets(SHEET_JOINTS)
        ws_summary = wb.Sheets(SHEET_SUMMARY)

        # Workbook-level inputs.
        ws_parts.Range("D1").Value = region       # PartFabRegion
        ws_parts.Range("F1").Value = length_uom    # PartLengthUnitsofMeasure

        _write_part_rows(ws_parts, clean_parts)
        _write_joint_rows(ws_joints, clean_joints)
        _write_summary_set_ids(ws_summary, clean_parts)

        # Recalc Steve's formulas natively.
        excel.Calculation = XL_CALC_AUTOMATIC
        excel.CalculateFull()

        summary = _read_outputs(ws_summary)
        part_costs = _read_part_costs(ws_parts)
        _scan_for_errors(wb, needs_review)

        wb.Save()
        return {
            "output_workbook": working_path,
            "summary": summary,
            "parts": part_costs,
            "needs_review": needs_review,
        }
    finally:
        # Guarantee no orphaned headless Excel process, even on error.
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        wb = None
        excel = None


# ---------------------------------------------------------------------------
# Input preparation + needs_review (no silent defaulting of required fields)
# ---------------------------------------------------------------------------
def _prepare_part(part: dict, needs_review: list[dict]) -> dict:
    pid = part.get("part_identifier", "<unknown>")
    p = dict(part)

    for field in REQUIRED_PART_FIELDS:
        v = p.get(field, None)
        if v is None or (isinstance(v, str) and v.strip() == ""):
            needs_review.append({
                "scope": "part", "identifier": pid, "field": field,
                "issue": "missing required input (not defaulted)",
            })

    # Cutting/forming may legitimately be defaulted (they are human dropdown picks).
    if not p.get("forming_method"):
        p["forming_method"] = FORMING_METHOD_DEFAULT
    if not p.get("cutting_method"):
        part_class = (p.get("part_class") or "").lower()
        if part_class == "tube":
            p["cutting_method"] = CUTTING_METHOD_DEFAULT_TUBE
        elif part_class == "sheet_metal":
            p["cutting_method"] = CUTTING_METHOD_DEFAULT_SHEET_METAL
        else:
            needs_review.append({
                "scope": "part", "identifier": pid, "field": "cutting_method",
                "issue": "no cutting method and unknown part_class; not defaulted",
            })
            p["cutting_method"] = None
    return p


def _prepare_joint(joint: dict, needs_review: list[dict]) -> dict:
    ident = f"{joint.get('part_a', '?')}-{joint.get('part_b', '?')}"
    for field in REQUIRED_JOINT_FIELDS:
        v = joint.get(field, None)
        if v is None or (isinstance(v, str) and v.strip() == ""):
            needs_review.append({
                "scope": "joint", "identifier": ident, "field": field,
                "issue": "missing required input (not defaulted)",
            })
    return dict(joint)


# ---------------------------------------------------------------------------
# Workbook staging + table writes
# ---------------------------------------------------------------------------
def _stage_workbook(template_path: str, output_dir: str) -> str:
    """Copy the pristine bundled template to a timestamped working file. The repo
    copy stays untouched; every run produces its own auditable workbook."""
    os.makedirs(output_dir, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    working = os.path.join(output_dir, f"CostEstimationOutput_{stamp}.xlsx")
    shutil.copyfile(template_path, working)
    return working


def _clear_table_rows(list_object) -> None:
    """Remove existing sample rows, preserving the table + its calculated-column
    definitions so appended rows re-acquire Steve's formulas."""
    body = list_object.DataBodyRange
    if body is not None:
        body.Delete()


def _write_part_rows(ws_parts, parts: list[dict]) -> None:
    lo = ws_parts.ListObjects(TABLE_PARTS)
    _clear_table_rows(lo)
    for p in parts:
        row = lo.ListRows.Add()          # appends + auto-fills calc columns
        cells = row.Range                # spans the whole table row
        for field, col in PART_COL.items():
            val = p.get(field, None)
            if val is not None:
                cells.Cells(1, col).Value = val


def _write_joint_rows(ws_joints, joints: list[dict]) -> None:
    lo = ws_joints.ListObjects(TABLE_JOINTS)
    _clear_table_rows(lo)
    for j in joints:
        row = lo.ListRows.Add()
        cells = row.Range
        for field, col in JOINT_COL.items():
            val = j.get(field, None)
            if val is not None:
                cells.Cells(1, col).Value = val
        # Weld Check (col O) deliberately left for the template's own formula.


def _write_summary_set_ids(ws_summary, parts: list[dict]) -> None:
    """One SummaryTable row per distinct part_set so the SUMIFS rollups resolve.
    Set ID goes in column A; the rest of the row is Steve's formulas."""
    seen, distinct = set(), []
    for p in parts:
        s = p.get("part_set")
        if s and s not in seen:
            seen.add(s)
            distinct.append(s)

    lo = ws_summary.ListObjects(TABLE_SUMMARY)
    _clear_table_rows(lo)
    for set_id in distinct:
        row = lo.ListRows.Add()
        row.Range.Cells(1, 1).Value = set_id   # Summary col A = Set ID


# ---------------------------------------------------------------------------
# Output read + error scan
# ---------------------------------------------------------------------------
def _read_outputs(ws_summary) -> list[dict]:
    """Read the recalculated SummaryTable into a list of {header: value} dicts."""
    lo = ws_summary.ListObjects(TABLE_SUMMARY)
    headers = [str(c.Value) for c in lo.HeaderRowRange.Cells]
    body = lo.DataBodyRange
    if body is None:
        return []

    raw = body.Value                     # tuple-of-tuples (or flat for 1 row)
    if raw is None:
        return []
    if not isinstance(raw[0], (tuple, list)):
        raw = (raw,)

    rows = []
    for r in raw:
        rows.append({headers[i]: r[i] for i in range(len(headers))})
    return rows


def _read_part_costs(ws_parts) -> list[dict]:
    """Read the recalculated per-part cost columns from PartsListTable.

    Returns one {column: value} dict per part, limited to PART_COST_COLUMNS
    (unit + extended material/labor/total cost), so the app can show a cost
    breakdown per part number without exposing the workbook's hundreds of
    intermediate calculation columns.
    """
    lo = ws_parts.ListObjects(TABLE_PARTS)
    headers = [str(c.Value) for c in lo.HeaderRowRange.Cells]
    body = lo.DataBodyRange
    if body is None:
        return []
    raw = body.Value
    if raw is None:
        return []
    if not isinstance(raw[0], (tuple, list)):
        raw = (raw,)

    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for r in raw:
        rows.append({col: r[idx[col]] for col in PART_COST_COLUMNS if col in idx})
    return rows


def _scan_for_errors(wb, needs_review: list[dict]) -> None:
    """Belt-and-suspenders: flag any Excel error cells left after recalc so a
    silent #REF!/#VALUE! never propagates into a dollar figure."""
    error_texts = {"#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}
    for sheet_name in (SHEET_PARTS, SHEET_JOINTS, SHEET_SUMMARY):
        ws = wb.Sheets(sheet_name)
        used = ws.UsedRange
        vals = used.Value
        if vals is None:
            continue
        if not isinstance(vals, tuple):     # single cell
            vals = ((vals,),)
        for row in vals:
            cells = row if isinstance(row, (tuple, list)) else (row,)
            for v in cells:
                if isinstance(v, str) and v in error_texts:
                    needs_review.append({
                        "scope": "workbook", "identifier": sheet_name,
                        "field": None, "issue": f"Excel error cell: {v}",
                    })
